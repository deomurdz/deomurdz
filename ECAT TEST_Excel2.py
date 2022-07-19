import pysoem
import time
import struct
import sys
import threading
import ctypes
import openpyxl
from collections import namedtuple


#Define Variables and Functions
pd_thread_stop_event = threading.Event()
ch_thread_stop_event = threading.Event()
actual_wkc = 0
EXCEL_SOURCE = "ECAT_Sine_Wave.xlsx"
ETHERNET_PORT = r'\Device\NPF_{90E40823-8B97-4C85-86D3-AB169710A76C}'

#Define Master
master = pysoem.Master()
master.in_op = False
master.do_check_state = False

#Define Functions

def processdata_thread():
    
    while not pd_thread_stop_event.is_set():
        #print("Sending")
        master.send_processdata()
        actual_wkc = master.receive_processdata(10000)
        if not actual_wkc == master.expected_wkc:
            print('Incorrect wkc')
        time.sleep(0.001)

        
def check_thread():

    while not ch_thread_stop_event.is_set():
        if master.in_op and ((actual_wkc < master.expected_wkc) or master.do_check_state):
            master.do_check_state = False
            master.read_state()
            for i, slave in enumerate(master.slaves):
                if slave.state != pysoem.OP_STATE:
                    master.do_check_state = True
                    #BasicExample._check_slave(slave, i) EXTRA FUNCTION NOT DEF
            if not master.do_check_state:
                print('OK\n')
        time.sleep(0.001)


def pdo_update_loop():

    master.in_op = True
    print("Entered PDO Loop")
    output_len = len(master.slaves[0].output)
    tmp = bytearray([0 for i in range(output_len)])
    toggle = True
    loop = True
    wb = openpyxl.load_workbook(EXCEL_SOURCE, data_only=True)
    sheet = wb.active
    MaxRow = sheet.max_row
    rowNum = 2
    
    try:
        while loop:
            if toggle:
                tmp[0] = 0x00
            else:
                Message = sheet.cell(row=rowNum,column=2).value
                if Message == "END":
                    rowNum = 2
                    Message = sheet.cell(row=rowNum,column=2).value
                Marr = bytes.fromhex(Message)
                print(Message)
                master.slaves[0].output = Marr
                rowNum += 1
                time.sleep(0.050)
            toggle ^= True
            
    except KeyboardInterrupt:
        #Ctrl-C abort handling
        loop = False
        print('***STOPPED***')





#Main
        
#Open PC Ethernet Port
master.open(ETHERNET_PORT)

#Find Slaves
if master.config_init() > 0:
    print("{} Slave Found and Configured".format(len(master.slaves)))
    slave = master.slaves[0]
    #device_bar = master.slaves[1]
else:
    print('No Device Found')

#Print Single Slave Name
print('    SLAVE 1: {}'.format(slave.name))

master.config_map()

#Wait 80 ms for all slaves to reach SAFE_OP state
if master.state_check(pysoem.SAFEOP_STATE, 80000) != pysoem.SAFEOP_STATE:
    master.read_state()
    for slave in master.slaves:
        if not slave.state == pysoem.SAFEOP_STATE:
            print('{} Did not reach SAFEOP state'.format(slave.name))
            print('Al status code {} ({})'.format(hex(slave.al_status),
                                                              pysoem.al_status_code_to_string(slave.al_status)))
    raise Exception('Not all slaves reached SAFEOP state')

#Test SDO Read
print("Expected VID: xxx")
readVID = struct.unpack('I', slave.sdo_read(0x1018, 1))[0]
print("  ACTUAL VID: "+hex(readVID))
slave.dc_sync(1, 10000000)
time.sleep(2)

#Attempt OP State
master.state = pysoem.OP_STATE
check_thread = threading.Thread(target=check_thread)
check_thread.start()
proc_thread = threading.Thread(target=processdata_thread)
proc_thread.start()
        
#Send One Valid PDO
master.send_processdata()
master.receive_processdata(2000)

#Request OP state for all slaves
master.write_state()
master.state_check(pysoem.OP_STATE, 50000)
all_slaves_reached_op_state = False
time.sleep(0.5)
print("Should be in OP, Wait 2 Seconds")
time.sleep(2)

#Check OP State
if master.state != pysoem.OP_STATE:
    master.read_state()
    for slave in master.slaves:
        if not slave.state == pysoem.OP_STATE:
            print('{} Did not reach OP state'.format(slave.name))
            print('Al status code {} ({})'.format(hex(slave.al_status),
                                                              pysoem.al_status_code_to_string(slave.al_status)))
    raise Exception('Not all slaves reached OP state')

all_slaves_reached_op_state = True
print("OPERATIONAL")
time.sleep(0.25)

#Run PDO Update Loop
if all_slaves_reached_op_state:
    pdo_update_loop()



#If Fail COM or CTRL+C To Stop
pd_thread_stop_event.set()
ch_thread_stop_event.set()
proc_thread.join()
check_thread.join()
master.state = pysoem.INIT_STATE

#Request INIT state for all slaves
master.write_state()

#Close Connection    
print("CLOSE")
master.close()
